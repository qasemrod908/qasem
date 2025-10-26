from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


def create_styled_workbook(title, headers, data, column_widths=None):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    row_colors = [
        PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ]
    
    for row_num, row_data in enumerate(data, 2):
        fill_color = row_colors[(row_num - 2) % 2]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = fill_color
    
    if column_widths:
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
    else:
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    ws.freeze_panes = 'A2'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_attendance_to_excel(attendance_records, filter_info=None):
    title = "سجل الحضور والغياب"
    
    headers = [
        "التاريخ",
        "الاسم الكامل",
        "النوع",
        "الحالة",
        "ملاحظات"
    ]
    
    data = []
    for record in attendance_records:
        user = record.user
        full_name = user.full_name if user else "غير محدد"
        user_type_display = "طالب" if record.user_type == "student" else "أستاذ" if record.user_type == "teacher" else record.user_type
        status_display = "حاضر" if record.status == "present" else "غائب" if record.status == "absent" else record.status
        
        data.append([
            record.date.strftime('%Y-%m-%d') if record.date else "",
            full_name,
            user_type_display,
            status_display,
            record.notes or ""
        ])
    
    column_widths = [15, 25, 15, 15, 30]
    
    return create_styled_workbook(title, headers, data, column_widths)


def export_students_to_excel(students):
    title = "قائمة الطلاب"
    
    headers = [
        "رقم الطالب",
        "الاسم الكامل",
        "رقم الهاتف",
        "الصف",
        "الشعبة",
        "اسم ولي الأمر",
        "رقم ولي الأمر",
        "الحالة",
        "تاريخ التسجيل"
    ]
    
    data = []
    for student in students:
        user = student.user
        status_display = "نشط" if user.is_active else "معطل"
        class_grade_name = student.class_grade.name if student.class_grade else "غير محدد"
        section_name = student.section.name if student.section else "غير محدد"
        
        data.append([
            student.student_number or "",
            user.full_name,
            student.phone or "",
            class_grade_name,
            section_name,
            student.guardian_name or "",
            student.guardian_phone or "",
            status_display,
            student.created_at.strftime('%Y-%m-%d') if student.created_at else ""
        ])
    
    column_widths = [15, 25, 15, 15, 15, 20, 15, 12, 15]
    
    return create_styled_workbook(title, headers, data, column_widths)


def export_teachers_to_excel(teachers):
    title = "قائمة الأساتذة"
    
    headers = [
        "الاسم الكامل",
        "رقم الهاتف",
        "التخصص",
        "سنوات الخبرة",
        "المؤهلات",
        "الحالة",
        "تاريخ التسجيل"
    ]
    
    data = []
    for teacher in teachers:
        user = teacher.user
        status_display = "نشط" if user.is_active else "معطل"
        
        data.append([
            user.full_name,
            teacher.phone or "",
            teacher.specialization or "",
            teacher.experience_years or 0,
            teacher.qualifications or "",
            status_display,
            teacher.created_at.strftime('%Y-%m-%d') if teacher.created_at else ""
        ])
    
    column_widths = [25, 15, 20, 15, 30, 12, 15]
    
    return create_styled_workbook(title, headers, data, column_widths)


def export_payments_to_excel(payments):
    title = "قائمة الأقساط والدفعات"
    
    headers = [
        "اسم الطالب",
        "عنوان القسط",
        "المبلغ الكلي",
        "المبلغ المدفوع",
        "المبلغ المتبقي",
        "الحالة",
        "تاريخ الاستحقاق",
        "تاريخ الإنشاء"
    ]
    
    data = []
    for payment in payments:
        student = payment.student
        student_name = student.user.full_name if student and student.user else "غير محدد"
        
        status_display = {
            'paid': 'مدفوع',
            'partial': 'مدفوع جزئياً',
            'pending': 'معلق'
        }.get(payment.status, payment.status)
        
        data.append([
            student_name,
            payment.title,
            f"{payment.total_amount:.2f}",
            f"{payment.paid_amount:.2f}",
            f"{payment.remaining_amount:.2f}",
            status_display,
            payment.due_date.strftime('%Y-%m-%d') if payment.due_date else "",
            payment.created_at.strftime('%Y-%m-%d') if payment.created_at else ""
        ])
    
    column_widths = [25, 25, 15, 15, 15, 15, 15, 15]
    
    return create_styled_workbook(title, headers, data, column_widths)
